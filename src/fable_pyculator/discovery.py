"""Workbook discovery helpers for public FABLE-C workbook conventions.

Discovery functions read workbook structure and return typed notebook declarations. They are
workbook-version informed, not a generic Excel conversion layer; generic extraction and formula
translation remain Modelwright responsibilities.
"""

from __future__ import annotations

from collections.abc import Iterable
from pathlib import Path
import re

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils.cell import get_column_letter, range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from fable_pyculator.spec import (
    FABLE_OUTPUT_SURFACE_SHEETS,
    HeadlinePoint,
    HeadlineSeries,
    OutputTable,
    ScenarioDefinitionTable,
    ScenarioParameter,
    SelectionControl,
    SelectionOption,
)


SCENARIO_SHEET_HINTS = ("scenario", "scenarios")
INPUT_LABEL_HINTS = ("scenario", "target", "assumption", "parameter", "select", "choice")
OUTPUT_COLUMN_FLAVOUR_TAG_PATTERN = re.compile(
    r"^(AUX|CALC|DIRECT|DATA\s*-\s*\d+(?:\.\d+)?|OUTPUT\s*-?\s*\d+(?:\s*,\s*\d+)*)$",
    re.IGNORECASE,
)
SCENARIO_DEFINITION_COLUMN_ROLE_TAG_PATTERN = re.compile(
    r"^(AUX|CALC|DIRECT|SCEN|DATA\s*-\s*\d+(?:\.\d+)?)$",
    re.IGNORECASE,
)
SCENARIO_DEFINITION_LOCATION_PATTERN = re.compile(r"^S\.\d+(?:\.[A-Z])?\.?$", re.IGNORECASE)


def discover_scenario_parameters(
    workbook_path: str | Path,
    *,
    sheet_hints: Iterable[str] = SCENARIO_SHEET_HINTS,
    label_hints: Iterable[str] = INPUT_LABEL_HINTS,
    max_rows: int = 250,
    max_columns: int = 40,
) -> list[ScenarioParameter]:
    """Return likely scenario controls from visible FABLE Calculator sheets.

    This is deliberately heuristic. It finds non-formula cells near text labels on sheets whose names
    look scenario-related; country-specific wrappers should review and curate the result into a
    committed spec before treating it as a stable user interface.
    """

    workbook = load_workbook(workbook_path, data_only=False, read_only=True)
    lowered_sheet_hints = tuple(hint.casefold() for hint in sheet_hints)
    lowered_label_hints = tuple(hint.casefold() for hint in label_hints)
    parameters: list[ScenarioParameter] = []

    for worksheet in workbook.worksheets:
        if not any(hint in worksheet.title.casefold() for hint in lowered_sheet_hints):
            continue
        for row in worksheet.iter_rows(max_row=max_rows, max_col=max_columns):
            label_cell = _first_label_cell(row, lowered_label_hints)
            if label_cell is None:
                continue
            label, label_column = label_cell
            value_cells = [cell for cell in row if getattr(cell, "column", 0) > label_column]
            if not value_cells:
                value_cells = list(row)
            for cell in value_cells:
                if _is_editable_value_cell(cell):
                    parameters.append(
                        ScenarioParameter(
                            name=_parameter_name(worksheet.title, cell.coordinate),
                            label=label,
                            cell_ref=f"{worksheet.title}!{cell.coordinate}",
                            kind=_control_kind(cell.value),
                            default=cell.value,
                            source="heuristic",
                        )
                    )
                    break
    return parameters


def discover_selection_controls(
    workbook_path: str | Path,
    *,
    sheet_name: str = "SCENARIOS selection",
) -> list[SelectionControl]:
    """Discover mutually exclusive ``x`` selection tables on ``SCENARIOS selection``.

    The public 2020 and 2021 FABLE-C workbooks expose 16 high-level scenario controls with this
    structure. The first table column is the marker column; the second column contains the option
    value passed to :meth:`fable_pyculator.SelectionControl.input_mapping`.
    """

    workbook = load_workbook(workbook_path, data_only=False, read_only=False)
    worksheet = workbook[sheet_name]
    controls: list[SelectionControl] = []
    for table_name in worksheet.tables.keys():
        table = worksheet.tables[table_name]
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        headers = [worksheet.cell(min_row, column).value for column in range(min_col, max_col + 1)]
        if len(headers) < 2 or str(headers[0]).casefold() != "selection":
            continue
        code_header = str(headers[1])
        label = _control_label(worksheet.cell(min_row - 1, min_col + 1).value, table_name)
        location = _optional_text(worksheet.cell(min_row - 1, min_col).value)
        options: list[SelectionOption] = []
        for row in range(min_row + 1, max_row + 1):
            value = worksheet.cell(row, min_col + 1).value
            if value is None:
                continue
            selected_marker = worksheet.cell(row, min_col).value
            options.append(
                SelectionOption(
                    value=str(value),
                    label=str(value),
                    selection_cell_ref=f"{worksheet.title}!{worksheet.cell(row, min_col).coordinate}",
                    description=_optional_text(worksheet.cell(row, min_col + 2).value),
                    selected=isinstance(selected_marker, str) and selected_marker.strip().casefold() == "x",
                )
            )
        controls.append(
            SelectionControl(
                name=_parameter_name("", table_name),
                label=label,
                table_name=table_name,
                sheet=worksheet.title,
                range_ref=table.ref,
                code_header=code_header,
                options=options,
                location=location,
            )
        )
    return sorted(controls, key=lambda control: _location_sort_key(control.location, control.table_name))


def discover_scenario_definition_tables(
    workbook_path: str | Path,
    *,
    sheet_name: str = "SCENARIOS definition",
) -> list[ScenarioDefinitionTable]:
    """Discover native Excel tables on ``SCENARIOS definition``.

    Returned tables preserve headers, row labels, cell references, current workbook values,
    role/source markers, and scenario-definition location markers. They are intended for inspection
    and later curation; they are not yet automatically exposed as editable widgets.
    """

    workbook = load_workbook(workbook_path, data_only=False, read_only=False)
    if sheet_name not in workbook.sheetnames:
        return []
    worksheet = workbook[sheet_name]
    tables: list[ScenarioDefinitionTable] = []
    for table_name in worksheet.tables.keys():
        table = worksheet.tables[table_name]
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        column_labels = tuple(
            _column_label(worksheet.cell(min_row, column).value, column)
            for column in range(min_col, max_col + 1)
        )
        role_tags, raw_role_tags, role_tag_refs = _scenario_definition_column_role_tags(
            worksheet,
            min_col,
            min_row,
            max_col,
        )
        locations, location_refs = _scenario_definition_location_markers(
            worksheet,
            min_col,
            min_row,
            max_col,
        )
        row_labels = tuple(
            _row_label(worksheet.cell(row, min_col).value, row)
            for row in range(min_row + 1, max_row + 1)
        )
        cell_refs = tuple(
            tuple(
                f"{worksheet.title}!{get_column_letter(column)}{row}"
                for column in range(min_col, max_col + 1)
            )
            for row in range(min_row + 1, max_row + 1)
        )
        values = tuple(
            tuple(worksheet.cell(row, column).value for column in range(min_col, max_col + 1))
            for row in range(min_row + 1, max_row + 1)
        )
        tables.append(
            ScenarioDefinitionTable(
                name=_parameter_name(sheet_name, table_name),
                sheet=worksheet.title,
                range_ref=table.ref,
                cell_refs=cell_refs,
                row_labels=row_labels,
                column_labels=column_labels,
                values=values,
                column_role_tags=role_tags,
                raw_column_role_tags=raw_role_tags,
                column_role_tag_refs=role_tag_refs,
                scenario_locations=locations,
                scenario_location_refs=location_refs,
                label=table_name,
            )
        )
    return sorted(tables, key=lambda definition_table: _range_sort_key(definition_table.range_ref))


def discover_output_tables(
    workbook_path: str | Path,
    *,
    sheet_names: Iterable[str] = FABLE_OUTPUT_SURFACE_SHEETS,
) -> list[OutputTable]:
    """Discover Excel tables on the canonical FABLE output data sheets.

    Output tables preserve workbook cell references and optional output-column flavour tags. The
    flavour metadata powers output DataFrame filtering in :func:`fable_pyculator.output_table_frame`.
    """

    workbook = load_workbook(workbook_path, data_only=False, read_only=False)
    tables: list[OutputTable] = []
    for sheet_name in sheet_names:
        if sheet_name not in workbook.sheetnames:
            continue
        worksheet = workbook[sheet_name]
        for table_name in worksheet.tables.keys():
            table = worksheet.tables[table_name]
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            column_labels = tuple(
                _column_label(worksheet.cell(min_row, column).value, column)
                for column in range(min_col, max_col + 1)
            )
            flavour_tags, raw_flavour_tags, flavour_tag_refs = _output_column_flavour_tags(
                worksheet,
                min_col,
                min_row,
                max_col,
            )
            row_labels = tuple(
                _row_label(worksheet.cell(row, min_col).value, row)
                for row in range(min_row + 1, max_row + 1)
            )
            cell_refs = tuple(
                tuple(
                    f"{worksheet.title}!{get_column_letter(column)}{row}"
                    for column in range(min_col, max_col + 1)
                )
                for row in range(min_row + 1, max_row + 1)
            )
            tables.append(
                OutputTable(
                    name=_parameter_name(sheet_name, table_name),
                    sheet=worksheet.title,
                    range_ref=table.ref,
                    cell_refs=cell_refs,
                    row_labels=row_labels,
                    column_labels=column_labels,
                    column_flavour_tags=flavour_tags,
                    raw_column_flavour_tags=raw_flavour_tags,
                    column_flavour_tag_refs=flavour_tag_refs,
                    label=table_name,
                )
            )
    return tables


def curate_default_headline_series(workbook_path: str | Path) -> list[HeadlineSeries]:
    """Curate the first FOOD, LAND, GHG, and WATER headline output series.

    The initial curation is intentionally narrow and provenance-friendly. It uses table descriptions
    from ``Indextables`` where available, then maps stable table columns on the canonical output
    sheets into notebook-ready time series.
    """

    workbook = load_workbook(workbook_path, data_only=False, read_only=False)
    descriptions = (
        _indextable_descriptions(workbook["Indextables"])
        if "Indextables" in workbook.sheetnames
        else {}
    )
    return [
        HeadlineSeries(
            name="food_total_kcal_feas",
            label="Feasible total kilocalorie consumption",
            group="FOOD",
            sheet="FOOD",
            table_name="Total_results_diets",
            points=_headline_points(
                workbook["FOOD"],
                "Total_results_diets",
                year_header="YEAR",
                value_headers=("kcal_feas",),
                row_filters={"PROD_GROUP": "TOTAL"},
            ),
            unit="kcal/cap/day",
            description=descriptions.get(_description_key("Total_results_diets")),
        ),
        HeadlineSeries(
            name="land_total_area",
            label="Total land area",
            group="LAND",
            sheet="LAND",
            table_name="ResultsLand",
            points=_headline_points(
                workbook["LAND"],
                "ResultsLand",
                year_header="Year",
                value_headers=("TOTAL",),
            ),
            description=descriptions.get(_description_key("ResultsLand")),
        ),
        HeadlineSeries(
            name="ghg_total_co2e",
            label="Total GHG emissions",
            group="GHG",
            sheet="GHG",
            table_name="ResultsGHG",
            points=_headline_points(
                workbook["GHG"],
                "ResultsGHG",
                year_header="Year",
                value_headers=("TotalCO2e",),
            ),
            description=descriptions.get(_description_key("ResultsGHG")),
        ),
        HeadlineSeries(
            name="water_total_footprint",
            label="Total water footprint",
            group="WATER",
            sheet="WATER",
            table_name="TotalResultsWF",
            points=_headline_points(
                workbook["WATER"],
                "TotalResultsWF",
                year_header="YEAR",
                value_headers=(
                    "wf_green_crop",
                    "wf_blue_crop",
                    "wf_grey_crop",
                    "wf_green_live",
                    "wf_blue_live",
                    "wf_grey_live",
                ),
                row_filters={"Product": "TOTAL"},
            ),
            description=descriptions.get(_description_key("TotalResultsWF")),
            aggregation="sum",
        ),
    ]


def _headline_points(
    worksheet: Worksheet,
    table_name: str,
    *,
    year_header: str,
    value_headers: tuple[str, ...],
    row_filters: dict[str, object] | None = None,
) -> tuple[HeadlinePoint, ...]:
    table = worksheet.tables[table_name]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    headers = {
        _optional_text(worksheet.cell(min_row, column).value): column
        for column in range(min_col, max_col + 1)
    }
    year_column = _required_header(headers, year_header, table_name)
    value_columns = tuple(_required_header(headers, header, table_name) for header in value_headers)
    filter_columns = {
        _required_header(headers, header, table_name): expected
        for header, expected in (row_filters or {}).items()
    }
    points: list[HeadlinePoint] = []
    for row in range(min_row + 1, max_row + 1):
        if any(worksheet.cell(row, column).value != expected for column, expected in filter_columns.items()):
            continue
        year = worksheet.cell(row, year_column).value
        if year is None:
            continue
        points.append(
            HeadlinePoint(
                year=year,
                cell_refs=tuple(
                    f"{worksheet.title}!{get_column_letter(column)}{row}"
                    for column in value_columns
                ),
            )
        )
    return tuple(points)


def _required_header(headers: dict[str | None, int], header: str, table_name: str) -> int:
    if header not in headers:
        raise KeyError(f"table {table_name!r} does not contain header {header!r}")
    return headers[header]


def _indextable_descriptions(worksheet: Worksheet) -> dict[str, str]:
    header_cells = next(worksheet.iter_rows(min_row=1, max_row=1), ())
    headers = [_optional_text(cell.value) for cell in header_cells]
    table_column = _header_index(headers, ("table", "name"))
    description_column = _header_index(headers, ("description", "table description"))
    if table_column is None or description_column is None:
        return {}
    descriptions: dict[str, str] = {}
    for row in worksheet.iter_rows(min_row=2):
        table_name = _optional_text(row[table_column].value)
        description = _optional_text(row[description_column].value)
        if table_name and description:
            descriptions[_description_key(table_name)] = description
    return descriptions


def _header_index(headers: list[str | None], candidates: tuple[str, ...]) -> int | None:
    lowered_candidates = {candidate.casefold() for candidate in candidates}
    for index, header in enumerate(headers):
        if header is not None and header.casefold() in lowered_candidates:
            return index
    return None


def _description_key(table_name: str) -> str:
    return "".join(character.casefold() for character in table_name if character.isalnum())


def _first_label_cell(row: Iterable[Cell], label_hints: tuple[str, ...]) -> tuple[str, int] | None:
    for cell in row:
        if isinstance(cell.value, str):
            text = " ".join(cell.value.split())
            if text and any(hint in text.casefold() for hint in label_hints):
                return text, cell.column
    return None


def _is_editable_value_cell(cell: Cell) -> bool:
    if cell.value is None:
        return False
    if isinstance(cell.value, str) and cell.value.startswith("="):
        return False
    return isinstance(cell.value, str | int | float | bool)


def _control_kind(value: object) -> str:
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, int | float) and not isinstance(value, bool):
        return "number"
    return "text"


def _parameter_name(sheet: str, coordinate: str) -> str:
    stem = "".join(character.lower() if character.isalnum() else "_" for character in sheet)
    suffix = "".join(character.lower() if character.isalnum() else "_" for character in coordinate)
    name = "_".join(part for part in f"{stem}_{suffix}".split("_") if part)
    return name


def _control_label(value: object, fallback: str) -> str:
    text = _optional_text(value)
    return text or fallback


def _optional_text(value: object) -> str | None:
    if value is None:
        return None
    text = " ".join(str(value).split())
    return text or None


def _location_sort_key(location: str | None, fallback: str) -> tuple[int, str]:
    if location and location.startswith("S."):
        try:
            return int(location.removeprefix("S.").rstrip(".")), fallback
        except ValueError:
            pass
    return 999, fallback


def _range_sort_key(range_ref: str) -> tuple[int, int, str]:
    min_col, min_row, *_ = range_boundaries(range_ref)
    return min_row, min_col, range_ref


def _column_label(value: object, column: int) -> str:
    return _optional_text(value) or get_column_letter(column)


def _row_label(value: object, row: int) -> str:
    return _optional_text(value) or str(row)


def _output_column_flavour_tags(
    worksheet: Worksheet,
    min_col: int,
    min_row: int,
    max_col: int,
) -> tuple[tuple[str | None, ...], tuple[str | None, ...], tuple[str | None, ...]]:
    tag_row = _output_column_flavour_tag_row(worksheet, min_col, min_row, max_col)
    if tag_row is None:
        return (), (), ()
    raw_tags: list[str | None] = []
    tags: list[str | None] = []
    refs: list[str | None] = []
    for column in range(min_col, max_col + 1):
        cell = worksheet.cell(tag_row, column)
        raw_tag = _optional_text(cell.value)
        raw_tags.append(raw_tag)
        tags.append(_canonical_output_column_flavour_tag(raw_tag))
        refs.append(f"{worksheet.title}!{cell.coordinate}")
    return tuple(tags), tuple(raw_tags), tuple(refs)


def _scenario_definition_column_role_tags(
    worksheet: Worksheet,
    min_col: int,
    min_row: int,
    max_col: int,
) -> tuple[tuple[str | None, ...], tuple[str | None, ...], tuple[str | None, ...]]:
    tag_row = _scenario_definition_column_role_tag_row(worksheet, min_col, min_row, max_col)
    if tag_row is None:
        return (), (), ()
    raw_tags: list[str | None] = []
    tags: list[str | None] = []
    refs: list[str | None] = []
    for column in range(min_col, max_col + 1):
        cell = worksheet.cell(tag_row, column)
        raw_tag = _optional_text(cell.value)
        raw_tags.append(raw_tag)
        tags.append(_canonical_scenario_definition_column_role_tag(raw_tag))
        refs.append(f"{worksheet.title}!{cell.coordinate}")
    return tuple(tags), tuple(raw_tags), tuple(refs)


def _scenario_definition_column_role_tag_row(
    worksheet: Worksheet,
    min_col: int,
    min_row: int,
    max_col: int,
) -> int | None:
    for row in range(min_row - 1, max(0, min_row - 10), -1):
        values = [worksheet.cell(row, column).value for column in range(min_col, max_col + 1)]
        if any(_canonical_scenario_definition_column_role_tag(_optional_text(value)) for value in values):
            return row
    return None


def _scenario_definition_location_markers(
    worksheet: Worksheet,
    min_col: int,
    min_row: int,
    max_col: int,
) -> tuple[tuple[str, ...], tuple[str, ...]]:
    locations: list[str] = []
    refs: list[str] = []
    for row in range(max(1, min_row - 10), min_row):
        for column in range(min_col, max_col + 1):
            cell = worksheet.cell(row, column)
            location = _canonical_scenario_definition_location(_optional_text(cell.value))
            if location is not None and location not in locations:
                locations.append(location)
                refs.append(f"{worksheet.title}!{cell.coordinate}")
    return tuple(locations), tuple(refs)


def _output_column_flavour_tag_row(
    worksheet: Worksheet,
    min_col: int,
    min_row: int,
    max_col: int,
) -> int | None:
    for row in range(min_row - 1, max(0, min_row - 10), -1):
        values = [worksheet.cell(row, column).value for column in range(min_col, max_col + 1)]
        if any(_canonical_output_column_flavour_tag(_optional_text(value)) for value in values):
            return row
    return None


def _canonical_output_column_flavour_tag(value: str | None) -> str | None:
    if value is None:
        return None
    text = re.sub(r"\s+", " ", value.strip()).upper()
    if not text:
        return None
    text = re.sub(r"^(DATA|OUTPUT)\s*-\s*", r"\1-", text)
    text = re.sub(r"^OUTPUT\s+(\d)", r"OUTPUT-\1", text)
    text = re.sub(r"\s*,\s*", ",", text)
    if OUTPUT_COLUMN_FLAVOUR_TAG_PATTERN.match(text):
        return text
    return None


def _canonical_scenario_definition_column_role_tag(value: str | None) -> str | None:
    if value is None:
        return None
    text = re.sub(r"\s+", " ", value.strip()).upper()
    if not text:
        return None
    text = re.sub(r"^DATA\s*-\s*", "DATA-", text)
    if SCENARIO_DEFINITION_COLUMN_ROLE_TAG_PATTERN.match(text):
        return text
    return None


def _canonical_scenario_definition_location(value: str | None) -> str | None:
    if value is None:
        return None
    text = re.sub(r"\s+", "", value.strip()).upper()
    if not text:
        return None
    if SCENARIO_DEFINITION_LOCATION_PATTERN.match(text):
        return text.rstrip(".")
    return None
