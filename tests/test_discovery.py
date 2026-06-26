from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.table import Table

from fable_pyculator import (
    curate_default_headline_series,
    discover_output_tables,
    discover_scenario_definition_tables,
    discover_scenario_parameters,
    discover_selection_controls,
)


def test_discover_scenario_parameters_finds_labeled_values(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "SCENARIOS selection"
    worksheet["A1"] = "Scenario ambition"
    worksheet["D1"] = 2
    worksheet["A2"] = "Formula output"
    worksheet["D2"] = "=D1*2"
    path = tmp_path / "fable.xlsx"
    workbook.save(path)

    parameters = discover_scenario_parameters(path)

    assert len(parameters) == 1
    assert parameters[0].cell_ref == "SCENARIOS selection!D1"
    assert parameters[0].default == 2


def test_discover_selection_controls_finds_x_marker_table(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "SCENARIOS selection"
    worksheet["A1"] = "S.1"
    worksheet["B1"] = "GDP projections"
    worksheet.append(["SELECTION", "GDP_SCEN", "DESCRIPTION"])
    worksheet.append([None, "SSP1", "Sustainability"])
    worksheet.append(["x", "SSP2", "Middle of the road"])
    worksheet.append([None, "SSP3", "Fragmentation"])
    worksheet.add_table(Table(displayName="GDP_Scen", ref="A2:C5"))
    path = tmp_path / "fable.xlsx"
    workbook.save(path)

    controls = discover_selection_controls(path)

    assert len(controls) == 1
    control = controls[0]
    assert control.name == "gdp_scen"
    assert control.label == "GDP projections"
    assert control.location == "S.1"
    assert control.default == "SSP2"
    assert control.input_mapping("SSP3") == {
        "SCENARIOS selection!A3": None,
        "SCENARIOS selection!A4": None,
        "SCENARIOS selection!A5": "x",
    }


def test_discover_output_tables_finds_tables_on_canonical_output_sheets(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "FOOD"
    worksheet.append(["DIRECT", "DATA -1", "OUTPUT - 8"])
    worksheet.append(["Metric", "2020", "2030"])
    worksheet.append(["Calories", 2500, 2600])
    worksheet.append(["Protein", 80, 82])
    worksheet.add_table(Table(displayName="Results_Diets", ref="A2:C4"))
    path = tmp_path / "fable.xlsx"
    workbook.save(path)

    tables = discover_output_tables(path)

    assert len(tables) == 1
    table = tables[0]
    assert table.name == "food_results_diets"
    assert table.sheet == "FOOD"
    assert table.range_ref == "A2:C4"
    assert table.column_labels == ("Metric", "2020", "2030")
    assert table.column_flavour_tags == ("DIRECT", "DATA-1", "OUTPUT-8")
    assert table.raw_column_flavour_tags == ("DIRECT", "DATA -1", "OUTPUT - 8")
    assert table.column_flavour_tag_refs == ("FOOD!A1", "FOOD!B1", "FOOD!C1")
    assert table.row_labels == ("Calories", "Protein")
    assert table.cell_refs == (
        ("FOOD!A3", "FOOD!B3", "FOOD!C3"),
        ("FOOD!A4", "FOOD!B4", "FOOD!C4"),
    )


def test_discover_scenario_definition_tables_preserves_table_values_and_tags(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "SCENARIOS definition"
    worksheet.append(["S.3.c", None, None, None])
    worksheet.append(["DIRECT", "SCEN", "DATA - 1", "CALC"])
    worksheet.append(["DietScen", "PROD_GROUP", "target", "diff"])
    worksheet.append(["Current", "CEREALS", 2500, "=C3-2400"])
    worksheet.append(["Ambitious", "CEREALS", 2600, "=C4-2400"])
    worksheet.add_table(Table(displayName="DietTarget", ref="A3:D5"))
    path = tmp_path / "fable.xlsx"
    workbook.save(path)

    tables = discover_scenario_definition_tables(path)

    assert len(tables) == 1
    table = tables[0]
    assert table.name == "scenarios_definition_diettarget"
    assert table.label == "DietTarget"
    assert table.sheet == "SCENARIOS definition"
    assert table.range_ref == "A3:D5"
    assert table.column_labels == ("DietScen", "PROD_GROUP", "target", "diff")
    assert table.column_role_tags == ("DIRECT", "SCEN", "DATA-1", "CALC")
    assert table.raw_column_role_tags == ("DIRECT", "SCEN", "DATA - 1", "CALC")
    assert table.column_role_tag_refs == (
        "SCENARIOS definition!A2",
        "SCENARIOS definition!B2",
        "SCENARIOS definition!C2",
        "SCENARIOS definition!D2",
    )
    assert table.scenario_locations == ("S.3.C",)
    assert table.scenario_location_refs == ("SCENARIOS definition!A1",)
    assert table.row_labels == ("Current", "Ambitious")
    assert table.cell_refs[0] == (
        "SCENARIOS definition!A4",
        "SCENARIOS definition!B4",
        "SCENARIOS definition!C4",
        "SCENARIOS definition!D4",
    )
    assert table.values[1] == ("Ambitious", "CEREALS", 2600, "=C4-2400")


def test_curate_default_headline_series_maps_core_output_tables(tmp_path: Path) -> None:
    workbook = Workbook()
    workbook.active.title = "Indextables"
    indextables = workbook["Indextables"]
    indextables.append(["Table", "Description"])
    indextables.append(["Total_Results_diets", "Diet totals"])
    indextables.append(["ResultsLand", "Land totals"])
    indextables.append(["ResultsGHG", "GHG totals"])
    indextables.append(["TotalResultsWF", "Water totals"])

    food = workbook.create_sheet("FOOD")
    food.append(["PROD_GROUP", "YEAR", "kcal_feas"])
    food.append(["TOTAL", 2030, 2500])
    food.append(["TOTAL", 2050, 2600])
    food.add_table(Table(displayName="Total_results_diets", ref="A1:C3"))

    land = workbook.create_sheet("LAND")
    land.append(["Year", "TOTAL"])
    land.append([2030, 100])
    land.add_table(Table(displayName="ResultsLand", ref="A1:B2"))

    ghg = workbook.create_sheet("GHG")
    ghg.append(["Year", "TotalCO2e"])
    ghg.append([2030, 42])
    ghg.add_table(Table(displayName="ResultsGHG", ref="A1:B2"))

    water = workbook.create_sheet("WATER")
    water.append(
        [
            "Product",
            "YEAR",
            "wf_green_crop",
            "wf_blue_crop",
            "wf_grey_crop",
            "wf_green_live",
            "wf_blue_live",
            "wf_grey_live",
        ]
    )
    water.append(["TOTAL", 2030, 1, 2, 3, 4, 5, 6])
    water.add_table(Table(displayName="TotalResultsWF", ref="A1:H2"))

    path = tmp_path / "fable.xlsx"
    workbook.save(path)

    series = curate_default_headline_series(path)

    assert [item.name for item in series] == [
        "food_total_kcal_feas",
        "land_total_area",
        "ghg_total_co2e",
        "water_total_footprint",
    ]
    assert series[0].description == "Diet totals"
    assert series[0].points[0].cell_refs == ("FOOD!C2",)
    assert series[3].aggregation == "sum"
    assert series[3].points[0].cell_refs == (
        "WATER!C2",
        "WATER!D2",
        "WATER!E2",
        "WATER!F2",
        "WATER!G2",
        "WATER!H2",
    )
