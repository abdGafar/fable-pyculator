from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.table import Table

from fable_pyculator import (
    build_2020_notebook_spec,
    load_generated_model,
    run_2020_notebook_loop,
    run_notebook_loop,
)


def test_load_generated_model_imports_calculate_module(tmp_path: Path) -> None:
    model_path = tmp_path / "generated_fable.py"
    model_path.write_text(
        "def calculate(inputs=None):\n"
        "    return {'GHG!B2': 42, 'input_count': len(inputs or {})}\n",
        encoding="utf-8",
    )

    module = load_generated_model(model_path, module_name="test_generated_fable")

    assert module.calculate({"SCENARIOS selection!A3": "x"}) == {"GHG!B2": 42, "input_count": 1}


def test_load_generated_model_reports_missing_path(tmp_path: Path) -> None:
    with pytest.raises(FileNotFoundError, match="generated model not found"):
        load_generated_model(tmp_path / "missing.py")


def test_run_notebook_loop_applies_selection_controls_and_renders_artifacts() -> None:
    spec = build_2020_notebook_spec(_synthetic_workbook_path())

    def calculate(inputs=None):
        assert inputs == {
            "SCENARIOS selection!A3": "x",
            "SCENARIOS selection!A4": None,
        }
        return {
            "GHG!A3": 2030,
            "GHG!B3": 42,
            "FOOD!C3": 2500,
            "LAND!B3": 100,
            "WATER!C3": 1,
            "WATER!D3": 2,
            "WATER!E3": 3,
            "WATER!F3": 4,
            "WATER!G3": 5,
            "WATER!H3": 6,
        }

    result = run_notebook_loop(
        calculate,
        spec,
        {"gdp_scen": "SSP1"},
        include_figures=False,
    )

    assert result.run.inputs == {
        "SCENARIOS selection!A3": "x",
        "SCENARIOS selection!A4": None,
    }
    assert set(result.output_tables) == {
        "food_total_results_diets",
        "land_resultsland",
        "ghg_resultsghg",
        "water_totalresultswf",
    }
    assert set(result.headline_frames) == {
        "food_total_kcal_feas",
        "land_total_area",
        "ghg_total_co2e",
        "water_total_footprint",
    }
    assert result.output_tables["ghg_resultsghg"].loc["2030", "TotalCO2e"] == 42
    assert result.headline_frames["ghg_total_co2e"].loc[0, "value"] == 42
    assert result.headline_frames["water_total_footprint"].loc[0, "value"] == 21
    assert result.headline_figures == {}


def test_build_2020_notebook_spec_includes_scenario_definition_tables() -> None:
    spec = build_2020_notebook_spec(_synthetic_workbook_path())

    assert [table.name for table in spec.scenario_definition_tables] == [
        "scenarios_definition_diettarget"
    ]
    assert spec.scenario_definition_tables[0].column_flavour_tags == ("DIRECT", "SCEN", "DATA-1")


def test_run_notebook_loop_preserves_explicit_rendered_artifact_subsets() -> None:
    spec = build_2020_notebook_spec(_synthetic_workbook_path())

    result = run_notebook_loop(
        lambda inputs=None: {
            "GHG!A3": 2030,
            "GHG!B3": 42,
            "WATER!C3": 1,
            "WATER!D3": 2,
            "WATER!E3": 3,
            "WATER!F3": 4,
            "WATER!G3": 5,
            "WATER!H3": 6,
        },
        spec,
        output_table_names=("ghg_resultsghg",),
        headline_series_names=("ghg_total_co2e",),
        include_figures=False,
    )

    assert set(result.output_tables) == {"ghg_resultsghg"}
    assert set(result.headline_frames) == {"ghg_total_co2e"}
    assert result.output_tables["ghg_resultsghg"].loc["2030", "TotalCO2e"] == 42


def test_run_2020_notebook_loop_loads_ignored_model_path(tmp_path: Path) -> None:
    workbook_path = _synthetic_workbook_path(tmp_path / "synthetic_fable.xlsx")
    model_path = tmp_path / "generated_fable_2020_model.py"
    model_path.write_text(
        "def calculate(inputs=None):\n"
        "    return {\n"
        "        'GHG!A3': 2030,\n"
        "        'GHG!B3': 42,\n"
        "        'FOOD!C3': 2500,\n"
        "        'LAND!B3': 100,\n"
        "        'WATER!C3': 1,\n"
        "        'WATER!D3': 2,\n"
        "        'WATER!E3': 3,\n"
        "        'WATER!F3': 4,\n"
        "        'WATER!G3': 5,\n"
        "        'WATER!H3': 6,\n"
        "    }\n",
        encoding="utf-8",
    )

    result = run_2020_notebook_loop(
        {"gdp_scen": "SSP2"},
        workbook_path=workbook_path,
        generated_model_path=model_path,
        output_table_names=("ghg_resultsghg",),
        output_table_column_flavour_tags="OUTPUT-8",
        headline_series_names=("ghg_total_co2e",),
        include_figures=False,
    )

    assert result.run.inputs == {
        "SCENARIOS selection!A3": None,
        "SCENARIOS selection!A4": "x",
    }
    assert result.headline_frames["ghg_total_co2e"].loc[0, "value"] == 42
    assert list(result.output_tables["ghg_resultsghg"].columns) == ["Year", "TotalCO2e"]


def _synthetic_workbook_path(path: Path | None = None) -> Path:
    workbook = Workbook()
    scenarios = workbook.active
    scenarios.title = "SCENARIOS selection"
    scenarios["A1"] = "S.1"
    scenarios["B1"] = "GDP projections"
    scenarios.append(["SELECTION", "GDP_SCEN", "DESCRIPTION"])
    scenarios.append([None, "SSP1", "Sustainability"])
    scenarios.append(["x", "SSP2", "Middle of the road"])
    scenarios.add_table(Table(displayName="GDP_Scen", ref="A2:C4"))

    definition = workbook.create_sheet("SCENARIOS definition")
    definition.append(["DIRECT", "SCEN", "DATA-1"])
    definition.append(["DietScen", "PROD_GROUP", "target"])
    definition.append(["Current", "CEREALS", 2500])
    definition.add_table(Table(displayName="DietTarget", ref="A2:C3"))

    indextables = workbook.create_sheet("Indextables")
    indextables.append(["Table", "Description"])
    indextables.append(["Total_Results_diets", "Diet totals"])
    indextables.append(["ResultsLand", "Land totals"])
    indextables.append(["ResultsGHG", "GHG totals"])
    indextables.append(["TotalResultsWF", "Water totals"])

    food = workbook.create_sheet("FOOD")
    food.append(["DIRECT", "DIRECT", "OUTPUT-7"])
    food.append(["PROD_GROUP", "YEAR", "kcal_feas"])
    food.append(["TOTAL", 2030, 2500])
    food.add_table(Table(displayName="Total_results_diets", ref="A2:C3"))

    land = workbook.create_sheet("LAND")
    land.append(["DIRECT", "OUTPUT-4"])
    land.append(["Year", "TOTAL"])
    land.append([2030, 100])
    land.add_table(Table(displayName="ResultsLand", ref="A2:B3"))

    ghg = workbook.create_sheet("GHG")
    ghg.append(["DIRECT", "OUTPUT - 8"])
    ghg.append(["Year", "TotalCO2e"])
    ghg.append([2030, 42])
    ghg.add_table(Table(displayName="ResultsGHG", ref="A2:B3"))

    water = workbook.create_sheet("WATER")
    water.append(["DIRECT", "DIRECT", "OUTPUT-9", "OUTPUT-9", "OUTPUT-9", "OUTPUT-9", "OUTPUT-9", "OUTPUT-9"])
    water.append(
        [
            "Product",
            "YEAR",
            "wf_green_crop",
            "wf_blue_crop",
            "wf_grey_crop",
            "wf_green_live",
            "wf_blue_live",
            "wf_grey_live",
        ]
    )
    water.append(["TOTAL", 2030, 1, 2, 3, 4, 5, 6])
    water.add_table(Table(displayName="TotalResultsWF", ref="A2:H3"))

    target = path or Path("tmp/test-scratch/synthetic_fable.xlsx")
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target)
    return target
